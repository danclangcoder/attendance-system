from datetime import datetime
from pathlib import Path
from openpyxl import Workbook as XLWorkbook, load_workbook
from db.database import get_logs

class Excel:
    def __init__(self, file_path: Path | None):
        self.file_path = file_path
        self.wb = None
        self.sheet = None
        if file_path:
            self._initialize_workbook(file_path)

    def _initialize_workbook(self, file_path: Path):
        if file_path.exists():
            self.wb = load_workbook(file_path)
        else:
            self.wb = XLWorkbook()

        self.sheet = self.wb.active
        self._ensure_headers()
        self.save()

    def _ensure_headers(self):
        headers = ['Student Number', 'Complete Name', 'Timestamp']
        if self.sheet.max_row == 1 and all(
            cell.value is None for cell in self.sheet[1]
        ):
            for col, header in enumerate(headers, 1):
                self.sheet.cell(row=1, column=col, value=header)

    def log_attendance(self, student_number: str, complete_name: str):
        if not self.sheet:
            raise ValueError("Excel file not loaded.")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.sheet.append([student_number, complete_name, timestamp])
        self.save()

    def save(self):
        if self.wb and self.file_path:
            self.wb.save(self.file_path)

    def load_file(self, file_path: Path):
        self.file_path = file_path
        self._initialize_workbook(file_path)
        
    def sync_db_to_excel(self, session_tag="default_session"):
        logs = get_logs(session_tag)

        if not self.sheet:
            raise ValueError("Excel file not loaded.")

        for row in logs:
            self.sheet.append([row[0], "", row[1]])

        self.save()