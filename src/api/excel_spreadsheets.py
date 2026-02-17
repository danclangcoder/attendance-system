from openpyxl import Workbook as XLWorkbook, load_workbook
from datetime import datetime
from pathlib import Path

class LocalExcel:
    def __init__(self, file_path: Path):
        self.file_path = file_path

        if file_path and file_path.exists():
            self.wb = load_workbook(file_path)
            self.sheet = self.wb.active
        elif file_path:
            self.wb = XLWorkbook()
            self.sheet = self.wb.active
        else:
            self.wb = None
            self.sheet = None
            return

        # Ensure headers exist
        if self.sheet.max_row == 1 and all(cell.value is None for cell in self.sheet[1]):
            headers = ['Student Number', 'Complete Name', 'Timestamp']
            for col, header in enumerate(headers, 1):
                self.sheet.cell(row=1, column=col, value=header)
            self.save()

    def log_attendance(self, student_number: str, complete_name: str):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.sheet.append([student_number, complete_name, timestamp])
        self.save()

    def save(self):
        self.wb.save(self.file_path)

    def load_file(self, file_path: Path):
        self.file_path = file_path

        if file_path.exists():
            self.wb = load_workbook(file_path)
            self.sheet = self.wb.active
        else:
            self.wb = XLWorkbook()
            self.sheet = self.wb.active

        # Ensure headers exist
        if self.sheet.max_row == 1 and all(cell.value is None for cell in self.sheet[1]):
            headers = ['Student Number', 'Complete Name', 'Timestamp']
            for col, header in enumerate(headers, 1):
                self.sheet.cell(row=1, column=col, value=header)

            self.save()
